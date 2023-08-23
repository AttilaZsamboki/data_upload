from django.http import HttpResponse
import subprocess
from .utils.google_maps import get_street_view, get_street_view_url
from .pen.utils import update_adatlap_fields
import math
import codecs
from .utils.google_maps import calculate_distance
from django.core.files import File
from django.db import IntegrityError
from rest_framework import generics, viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.status import HTTP_201_CREATED, HTTP_200_OK, HTTP_400_BAD_REQUEST, HTTP_500_INTERNAL_SERVER_ERROR
from django.http import JsonResponse
import psycopg2
from . import models, serializers
from .permissions import AuthorAllUser
from rest_framework.decorators import api_view
import os
from io import open
from PIL import Image
from django.core.management import call_command
from datetime import date, timedelta
import json
from .sm.inventory_planner import inventory_planner
from openpyxl import load_workbook
from .utils.utils import log
from urllib.parse import quote


@api_view(["GET"])
def DownloadFile(request):
    if request.method == "GET":
        path_to_file = "/home/atti/googleds/files/" + request.GET.get("path")
        if os.path.exists(path_to_file):
            data = open(path_to_file, "rb")
            response = HttpResponse(
                File(data).read())
            response['Content-Disposition'] = 'attachment; filename=%s' % path_to_file.split(
                "/")[-1]
            return response

#----------------------------------------------------GENERIC-------------------------------------------------------#


@api_view(["POST"])
def CreateCashflowPlanner():
    try:
        DB_HOST = os.environ.get("DB_HOST")
        DB_NAME = os.environ.get("DB_NAME")
        DB_USER = os.environ.get("DB_USER")
        DB_PASS = os.environ.get("DB_PASS")
        DB_PORT = os.environ.get("DB_PORT")
        conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                                password=DB_PASS, host=DB_HOST, port=DB_PORT)
        cur = conn.cursor()

        start_date = date(2022, 10, 29)
        end_date = date(2024, 10, 29)

        delta = end_date - start_date

        days = []
        for i in range(delta.days + 1):
            day = start_date + timedelta(days=i)
            days.append(day)

        cur.execute(
            "select string_agg(distinct name, ', ') from cashflow_planner_table;")
        current_categories = list(cur.fetchone())[0].split(
            ", ")
        cur.execute('select distinct name, elem_tipus from elemek;')
        all_data = [list(i) for i in cur.fetchall()]
        all_categories = [i[0] for i in all_data]
        for i in current_categories:
            if i in all_categories:
                continue
            else:
                cur.execute(
                    f"delete from cashflow_planner_table where name = '{i}'")
                conn.commit()
        for i in all_data:
            if i[0] in current_categories:
                continue
            for day in days:
                cur.execute(
                    f"insert into cashflow_planner_table (name, day, planned_expense, tipus) values ('{i[0]}', '{day}', 0, '{i[1]}')")
                conn.commit()
        cur.close()
        conn.close()
    except:
        return HttpResponse("Bad")
    else:
        return HttpResponse("Good")


@api_view(["POST"])
def UploadProfileImg(request):
    if request.POST:
        data = request.data

        def path(type):
            return "/home/atti/googleds/dataupload/frontend/static/images/" + \
                str(data[f'{type}Image'])
        if os.path.exists(path("old")):
            os.replace(
                path("old"), path("new"))
            img = Image.open(data["newImage"])
            img.save(
                path("new"))
        else:
            img = Image.open(data["newImage"])
            img.save(
                path("new"))
        call_command("collectstatic", interactive=False)
        return HttpResponse("good")


@api_view(["GET"])
def UploadTimer(request):
    return HttpResponse(open("/home/atti/googleds/files/misc/last_upload.txt", "r"))


def ColumnNames(request):
    conn = psycopg2.connect(dbname="defaultdb", user="doadmin",
                            password="AVNS_FovmirLSFDui0KIAOnu", host="defaultdb.c0rzdkeutp8f.eu-central-1.rds.amazonaws.com", port=25060)
    if request.method == 'GET':
        cur = conn.cursor()
        cur.execute(
            "select table_name, column_name, data_type from information_schema.columns where table_schema = 'public'")
        return JsonResponse((cur.fetchall()), safe=False, json_dumps_params={'ensure_ascii': False})
    cur.close()
    conn.close()


def TableNames(request):
    conn = psycopg2.connect(dbname="defaultdb", user="doadmin",
                            password="AVNS_FovmirLSFDui0KIAOnu", host="defaultdb.c0rzdkeutp8f.eu-central-1.rds.amazonaws.com", port=25060)
    if request.method == 'GET':
        cur = conn.cursor()
        cur.execute(
            "select string_agg(table_name, ', ') from information_schema.tables where table_schema = 'public';")
        return JsonResponse(list(cur.fetchone())[0].split(", "), safe=False, json_dumps_params={'ensure_ascii': False})
    cur.close()
    conn.close()

# dataupload config


class TemplatesList(generics.ListCreateAPIView):
    queryset = models.DatauploadTabletemplates.objects.all()
    serializer_class = serializers.TemplatesSerializer
    # permission_classes = [AuthorAllUser]


class TemplateDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.DatauploadTabletemplates.objects.all()
    serializer_class = serializers.TemplatesSerializer
    permission_classes = [AuthorAllUser]


class UploadmodelList(generics.ListCreateAPIView):
    queryset = models.DatauploadUploadmodel.objects.all()
    serializer_class = serializers.UploadModelSerializer
    permission_classes = [AuthorAllUser]


class UploadmodelDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.DatauploadUploadmodel.objects.all()
    serializer_class = serializers.UploadModelSerializer
    permission_classes = [AuthorAllUser]


class TableOverviewDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.DatauploadTableOverview.objects.all()
    serializer_class = serializers.TableoverviewSerializer
    permission_classes = [AuthorAllUser]


class TableOverviewList(generics.ListCreateAPIView):
    queryset = models.DatauploadTableOverview.objects.all()
    serializer_class = serializers.TableoverviewSerializer
    permission_classes = [AuthorAllUser]


# -------------------------------------------------- DATAS --------------------------------------------------------------- #
    # ------------------------------------------------FOL-------------------------------------------------------------#

class FolCFDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolCF.objects.all()
    serializer_class = serializers.FolCFSerializer
    permission_classes = [AuthorAllUser]


class FolCFList(generics.ListCreateAPIView):
    queryset = models.FolCF.objects.all()
    serializer_class = serializers.FolCFSerializer
    permission_classes = [AuthorAllUser]


class FolFedezetDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolFedezet.objects.all()
    serializer_class = serializers.FolOrdersÖsszesítőSerializer
    permission_classes = [AuthorAllUser]


class FolOrdersÖsszesítőList(generics.ListCreateAPIView):
    queryset = models.FolOrdersÖsszesítő.objects.all()
    serializer_class = serializers.FolOrdersÖsszesítőSerializer
    permission_classes = [AuthorAllUser]


class FolOrdersÖsszesítőDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolOrdersÖsszesítő.objects.all()
    serializer_class = serializers.FolOrdersÖsszesítőSerializer
    permission_classes = [AuthorAllUser]


class FolFedezetList(generics.ListCreateAPIView):
    queryset = models.FolFedezet.objects.all()
    serializer_class = serializers.FolFedezetSerializer
    permission_classes = [AuthorAllUser]


class FolGlsOsszesitesDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolGlsOsszesites.objects.all()
    serializer_class = serializers.FolGlsOsszesitesSerializer
    permission_classes = [AuthorAllUser]


class FolGlsOsszesitesList(generics.ListCreateAPIView):
    queryset = models.FolGlsOsszesites.objects.all()
    serializer_class = serializers.FolGlsOsszesitesSerializer
    permission_classes = [AuthorAllUser]


class FolArresFigyeloDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolArresFigyelo.objects.all()
    serializer_class = serializers.FolArresFigyeloSerializer
    permission_classes = [AuthorAllUser]


class FolArresFigyeloList(generics.ListCreateAPIView):
    queryset = models.FolArresFigyelo.objects.all()
    serializer_class = serializers.FolArresFigyeloSerializer
    permission_classes = [AuthorAllUser]

# Fol Bevételek


class FolBevetelekList(generics.ListCreateAPIView):
    queryset = models.FolBevtelek.objects.all()
    serializer_class = serializers.FolBevetelekSerializer
    permission_classes = [AuthorAllUser]


class FolBevetelDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolBevtelek.objects.all()
    serializer_class = serializers.FolBevetelekSerializer
    permission_classes = [AuthorAllUser]


# Fol Költségek
class FolKoltsegekList(generics.ListCreateAPIView):
    queryset = models.FolKltsgek.objects.all()
    serializer_class = serializers.FolKoltsegekSerializer
    permission_classes = [AuthorAllUser]


class FolKoltsegDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolKltsgek.objects.all()
    serializer_class = serializers.FolKoltsegekSerializer
    permission_classes = [AuthorAllUser]

# Fol orders


class FolOrdersList(generics.ListCreateAPIView):
    queryset = models.FolOrders.objects.all()
    serializer_class = serializers.FolOrdersSerializer
    permission_classes = [AuthorAllUser]


class FolOrderDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolOrders.objects.all()
    serializer_class = serializers.FolOrdersSerializer
    permission_classes = [AuthorAllUser]

# Fol product suppliers


class FolProductSuppliersList(generics.ListCreateAPIView):
    queryset = models.FolProductSuppliers.objects.all()
    serializer_class = serializers.FolProductSuppliersSerializer
    permission_classes = [AuthorAllUser]


class FolProductSupplierDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolProductSuppliers.objects.all()
    serializer_class = serializers.FolProductSuppliersSerializer
    permission_classes = [AuthorAllUser]

# Fol stock report


class FolStockReportList(generics.ListCreateAPIView):
    queryset = models.FolStockReport.objects.all()
    serializer_class = serializers.FolStockReportSerializer
    permission_classes = [AuthorAllUser]


class FolStockReportDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolStockReport.objects.all()
    serializer_class = serializers.FolStockReportSerializer
    permission_classes = [AuthorAllUser]

# Fol számlák


class FolSzamlakList(generics.ListCreateAPIView):
    queryset = models.FolSzmlk.objects.all()
    serializer_class = serializers.FolSzamlakSerializer
    permission_classes = [AuthorAllUser]


class FolSzamlaDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolSzmlk.objects.all()
    serializer_class = serializers.FolSzamlakSerializer
    permission_classes = [AuthorAllUser]

# Fol unas


class FolUnasList(generics.ListCreateAPIView):
    queryset = models.FolUnas.objects.all()
    serializer_class = serializers.FolUnasSerializer
    permission_classes = [AuthorAllUser]


class FolUnasDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolUnas.objects.all()
    serializer_class = serializers.FolUnasSerializer
    permission_classes = [AuthorAllUser]


# Fol Gls elszámolás
class FolGlsElszámolásList(generics.ListCreateAPIView):
    queryset = models.FolGlsElszmols.objects.all()
    serializer_class = serializers.FolGlsElszámolsSerializer
    permission_classes = [AuthorAllUser]


class FolGlsElszámolásDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolGlsElszmols.objects.all()
    serializer_class = serializers.FolGlsElszámolsSerializer
    permission_classes = [AuthorAllUser]


class FolStockTransactionReportList(generics.ListCreateAPIView):
    queryset = models.FolStockTransactionReport.objects.all()
    serializer_class = serializers.FolStockTransactionReportSerializer
    permission_classes = [AuthorAllUser]


class FolStockTransactionReportDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolStockTransactionReport.objects.all()
    serializer_class = serializers.FolStockTransactionReportSerializer
    permission_classes = [AuthorAllUser]


class FolStockAgingList(generics.ListCreateAPIView):
    queryset = models.FolStockAging.objects.all()
    serializer_class = serializers.FolStockAgingSerializer
    permission_classes = [AuthorAllUser]


class FolReturnOrderShippingFeeList(generics.ListCreateAPIView):
    queryset = models.FolReturnOrderShippingFee.objects.all()
    serializer_class = serializers.FolReturnOrderShippingFeeSerializer
    permission_classes = [AuthorAllUser]


class FolReturnOrderItemList(generics.ListCreateAPIView):
    queryset = models.FolReturnOrderItem.objects.all()
    serializer_class = serializers.FolReturnOrderItemSerializer
    permission_classes = [AuthorAllUser]


class FolReturnOrderList(generics.ListCreateAPIView):
    queryset = models.FolReturnOrder.objects.all()
    serializer_class = serializers.FolReturnOrderSerializer
    permission_classes = [AuthorAllUser]


class FolOrderShippingFeeList(generics.ListCreateAPIView):
    queryset = models.FolReturnOrder.objects.all()
    serializer_class = serializers.FolReturnOrderSerializer
    permission_classes = [AuthorAllUser]


class FolOrderItemList(generics.ListCreateAPIView):
    queryset = models.FolOrderItem.objects.all()
    serializer_class = serializers.FolOrderItemSerializer
    permission_classes = [AuthorAllUser]


class FolOrderFeeList(generics.ListCreateAPIView):
    queryset = models.FolOrderFee.objects.all()
    serializer_class = serializers.FolOrderFeeSerializer
    permission_classes = [AuthorAllUser]


class FolOrderEtcList(generics.ListCreateAPIView):
    queryset = models.FolOrderEtc.objects.all()
    serializer_class = serializers.FolOrderEtcSerializer
    permission_classes = [AuthorAllUser]


class FolOrderBaseList(generics.ListCreateAPIView):
    queryset = models.FolOrderBase.objects.all()
    serializer_class = serializers.FolOrderBaseSerializer
    permission_classes = [AuthorAllUser]

#


class FolStockAgingDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolStockAging.objects.all()
    serializer_class = serializers.FolStockAgingSerializer
    permission_classes = [AuthorAllUser]


class FolReturnOrderShippingFeeDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolReturnOrderShippingFee.objects.all()
    serializer_class = serializers.FolReturnOrderShippingFeeSerializer
    permission_classes = [AuthorAllUser]


class FolReturnOrderItemDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolReturnOrderItem.objects.all()
    serializer_class = serializers.FolReturnOrderItemSerializer
    permission_classes = [AuthorAllUser]


class FolReturnOrderDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolReturnOrder.objects.all()
    serializer_class = serializers.FolReturnOrderSerializer
    permission_classes = [AuthorAllUser]


class FolOrderShippingFeeDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolReturnOrder.objects.all()
    serializer_class = serializers.FolReturnOrderSerializer
    permission_classes = [AuthorAllUser]


class FolOrderItemDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolOrderItem.objects.all()
    serializer_class = serializers.FolOrderItemSerializer
    permission_classes = [AuthorAllUser]


class FolOrderFeeDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolOrderFee.objects.all()
    serializer_class = serializers.FolOrderFeeSerializer
    permission_classes = [AuthorAllUser]


class FolOrderEtcDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolOrderEtc.objects.all()
    serializer_class = serializers.FolOrderEtcSerializer
    permission_classes = [AuthorAllUser]


class FolOrderBaseDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolOrderBase.objects.all()
    serializer_class = serializers.FolOrderBaseSerializer
    permission_classes = [AuthorAllUser]

    #------------------------------------------------PRO----------------------------------------------------------------#


class ProBevetelekList(generics.ListCreateAPIView):
    queryset = models.ProBevtelek.objects.all()
    serializer_class = serializers.ProBevetelekSerializer
    permission_classes = [AuthorAllUser]


class ProBevetelDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProBevtelek.objects.all()
    serializer_class = serializers.ProBevetelekSerializer
    permission_classes = [AuthorAllUser]


# Pro Költségek
class ProKoltsegekList(generics.ListCreateAPIView):
    queryset = models.ProKltsgek.objects.all()
    serializer_class = serializers.ProKoltsegekSerializer
    permission_classes = [AuthorAllUser]


class ProKoltsegDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProKltsgek.objects.all()
    serializer_class = serializers.ProKoltsegekSerializer
    permission_classes = [AuthorAllUser]

# Pro orders


class ProOrdersList(generics.ListCreateAPIView):
    queryset = models.ProOrders.objects.all()
    serializer_class = serializers.ProOrdersSerializer
    permission_classes = [AuthorAllUser]


class ProOrderDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProOrders.objects.all()
    serializer_class = serializers.ProOrdersSerializer
    permission_classes = [AuthorAllUser]

# Pro product suppliers


class ProProductSuppliersList(generics.ListCreateAPIView):
    queryset = models.ProProductSuppliers.objects.all()
    serializer_class = serializers.ProProductSuppliersSerializer
    permission_classes = [AuthorAllUser]


class ProProductSupplierDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProProductSuppliers.objects.all()
    serializer_class = serializers.ProProductSuppliersSerializer
    permission_classes = [AuthorAllUser]

# Pro stock report


class ProStockReportList(generics.ListCreateAPIView):
    queryset = models.ProStockReport.objects.all()
    serializer_class = serializers.ProStockReportSerializer
    permission_classes = [AuthorAllUser]


class ProStockReportDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProStockReport.objects.all()
    serializer_class = serializers.ProStockReportSerializer
    permission_classes = [AuthorAllUser]

# Pro számlák


class ProSzamlakList(generics.ListCreateAPIView):
    queryset = models.ProSzmlk.objects.all()
    serializer_class = serializers.ProSzamlakSerializer
    permission_classes = [AuthorAllUser]


class ProSzamlaDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProSzmlk.objects.all()
    serializer_class = serializers.ProSzamlakSerializer
    permission_classes = [AuthorAllUser]


class ProStockTransactionReportList(generics.ListCreateAPIView):
    queryset = models.ProStockTransactionReport.objects.all()
    serializer_class = serializers.ProStockTransactionReportSerializer
    permission_classes = [AuthorAllUser]


class ProStockTransactionReportDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProStockTransactionReport.objects.all()
    serializer_class = serializers.ProStockTransactionReportSerializer
    permission_classes = [AuthorAllUser]


class ProStockAgingList(generics.ListCreateAPIView):
    queryset = models.ProStockAging.objects.all()
    serializer_class = serializers.ProStockAgingSerializer
    permission_classes = [AuthorAllUser]


class ProReturnOrderShippingFeeList(generics.ListCreateAPIView):
    queryset = models.ProReturnOrderShippingFee.objects.all()
    serializer_class = serializers.ProReturnOrderShippingFeeSerializer
    permission_classes = [AuthorAllUser]


class ProReturnOrderItemList(generics.ListCreateAPIView):
    queryset = models.ProReturnOrderItem.objects.all()
    serializer_class = serializers.ProReturnOrderItemSerializer
    permission_classes = [AuthorAllUser]


class ProReturnOrderList(generics.ListCreateAPIView):
    queryset = models.ProReturnOrder.objects.all()
    serializer_class = serializers.ProReturnOrderSerializer
    permission_classes = [AuthorAllUser]


class ProOrderShippingFeeList(generics.ListCreateAPIView):
    queryset = models.ProReturnOrder.objects.all()
    serializer_class = serializers.ProReturnOrderSerializer
    permission_classes = [AuthorAllUser]


class ProOrderItemList(generics.ListCreateAPIView):
    queryset = models.ProOrderItem.objects.all()
    serializer_class = serializers.ProOrderItemSerializer
    permission_classes = [AuthorAllUser]


class ProOrderFeeList(generics.ListCreateAPIView):
    queryset = models.ProOrderFee.objects.all()
    serializer_class = serializers.ProOrderFeeSerializer
    permission_classes = [AuthorAllUser]


class ProOrderEtcList(generics.ListCreateAPIView):
    queryset = models.ProOrderEtc.objects.all()
    serializer_class = serializers.ProOrderEtcSerializer
    permission_classes = [AuthorAllUser]


class ProOrderBaseList(generics.ListCreateAPIView):
    queryset = models.ProOrderBase.objects.all()
    serializer_class = serializers.ProOrderBaseSerializer
    permission_classes = [AuthorAllUser]

#


class ProStockAgingDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProStockAging.objects.all()
    serializer_class = serializers.ProStockAgingSerializer
    permission_classes = [AuthorAllUser]


class ProReturnOrderShippingFeeDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProReturnOrderShippingFee.objects.all()
    serializer_class = serializers.ProReturnOrderShippingFeeSerializer
    permission_classes = [AuthorAllUser]


class ProReturnOrderItemDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProReturnOrderItem.objects.all()
    serializer_class = serializers.ProReturnOrderItemSerializer
    permission_classes = [AuthorAllUser]


class ProReturnOrderDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProReturnOrder.objects.all()
    serializer_class = serializers.ProReturnOrderSerializer
    permission_classes = [AuthorAllUser]


class ProOrderShippingFeeDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProReturnOrder.objects.all()
    serializer_class = serializers.ProReturnOrderSerializer
    permission_classes = [AuthorAllUser]


class ProOrderItemDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProOrderItem.objects.all()
    serializer_class = serializers.ProOrderItemSerializer
    permission_classes = [AuthorAllUser]


class ProOrderFeeDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProOrderFee.objects.all()
    serializer_class = serializers.ProOrderFeeSerializer
    permission_classes = [AuthorAllUser]


class ProOrderEtcDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProOrderEtc.objects.all()
    serializer_class = serializers.ProOrderEtcSerializer
    permission_classes = [AuthorAllUser]


class ProOrderBaseDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProOrderBase.objects.all()
    serializer_class = serializers.ProOrderBaseSerializer
    permission_classes = [AuthorAllUser]


class ProProductsDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProProducts.objects.all()
    serializer_class = serializers.ProProductsSerializer
    permission_classes = [AuthorAllUser]


class ProProductsList(generics.ListCreateAPIView):
    queryset = models.ProProducts.objects.all()
    serializer_class = serializers.ProProductsSerializer
    permission_classes = [AuthorAllUser]


class ProProductsDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.ProProducts.objects.all()
    serializer_class = serializers.ProProductsSerializer
    permission_classes = [AuthorAllUser]


class ProProductsList(generics.ListCreateAPIView):
    queryset = models.ProProducts.objects.all()
    serializer_class = serializers.ProProductsSerializer
    permission_classes = [AuthorAllUser]


class FolLearnDashDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.FolLearnDash.objects.all()
    serializer_class = serializers.FolLearnDashSerializer
    permission_classes = [AuthorAllUser]


class FolLearnDashList(generics.ListCreateAPIView):
    queryset = models.FolLearnDash.objects.all()
    serializer_class = serializers.FolLearnDashSerializer
    permission_classes = [AuthorAllUser]


class FeedDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.Feed.objects.all()
    serializer_class = serializers.FeedSerializer
    permission_classes = [AuthorAllUser]


class FeedList(generics.ListCreateAPIView):
    queryset = models.Feed.objects.all()
    serializer_class = serializers.FeedSerializer
    permission_classes = [AuthorAllUser]


class GroupsList(generics.ListCreateAPIView):
    queryset = models.DatauploadGroups.objects.all()
    serializer_class = serializers.GroupsSerializer
    permission_classes = [AuthorAllUser]


class GroupsDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.DatauploadGroups.objects.all()
    serializer_class = serializers.GroupsSerializer
    permission_classes = [AuthorAllUser]


class SMVendorDataSet(APIView):
    def get(self, request, format=None):
        filter_param = request.GET.get('filter', None)
        if filter_param:
            queryset = models.SMVendorData.objects.filter(is_visible=True)
        else:
            queryset = models.SMVendorData.objects.all()
        serializer_class = serializers.SMVendorDataSerializer(
            queryset, many=True)
        return Response(serializer_class.data)

    def post(self, request):
        data = json.loads(request.body)
        for item in data:
            if item["vendor"] is None:
                continue
            keys = ['vendor', 'lost_revenue', 'to_order_cost', 'replenish_date',
                    'to_order', 'inventory_value', 'latest_order_date', 'avg_lead_time', 'name']
            formatted_item = {k: v for k, v in item.items() if k not in keys}
            models.SMVendorsTable(
                name=item["vendor"], **formatted_item).save()
        return Response({'status': 'success'}, status=HTTP_201_CREATED)


class SMProductView(viewsets.ReadOnlyModelViewSet):
    queryset = models.SMProductView.objects.all()
    serializer_class = serializers.SMProductViewSerializer


class SMVendorOrders(APIView):
    def get(self, request, format=None):
        queryset = models.SMVendorOrders.objects.all()
        serializer_class = serializers.SMVendorOrdersSerializer(
            queryset, many=True)
        return Response(serializer_class.data)

    def post(self, request):
        vendor = json.loads(request.body)
        vendorObj = models.SMVendorsTable.objects.filter(
            name=vendor)
        if vendorObj:
            need_permission = vendorObj[0].need_permission
        else:
            need_permission = True
        status = ""
        if need_permission == True:
            status = "DRAFT"
        else:
            status = "OPEN"
        create_order = inventory_planner(vendor, status=status, is_new=True)
        if create_order["status"] == "ERROR":
            return Response({'status': 'failed', 'reason': create_order["message"]}, status=HTTP_400_BAD_REQUEST)
        return Response({'status': 'success'}, status=HTTP_201_CREATED)

    def put(self, request):
        data = json.loads(request.body)
        id, status = data["id"], data["status"]
        orderObj = models.SMVendorOrders.objects.filter(id=id)
        if orderObj:
            create_order = inventory_planner(orderObj[0].vendor,
                                             status=status, is_new=False, id=id)
            if create_order["status"] == "ERROR":
                return Response({'status': 'failed', 'reason': create_order["message"]}, status=HTTP_400_BAD_REQUEST)
            return Response({'status': 'success', 'message': create_order["message"]}, status=HTTP_200_OK)
        return Response({'status': 'failed', 'reason': f'null order object. id: {id}'}, status=HTTP_400_BAD_REQUEST)


class SMVendorOrdersDetail(APIView):
    def delete(self, request, id, format=None):
        orderObj = models.SMVendorOrders.objects.filter(id=id)
        if orderObj:
            orderObj = orderObj[0]
            if orderObj.order_status == "DRAFT":
                orderObj.delete()
                return Response({'status': 'success'}, status=HTTP_200_OK)
        return Response({'status': 'failed', 'reason': f'null order object. id: {id}'}, status=HTTP_400_BAD_REQUEST)


class ExcelFileView(APIView):

    def get(self, request, vendor, date, format=None):
        workbook = load_workbook(
            filename=f"/home/atti/googleds/files/sm_pos/{vendor}/{date}.xlsx")
        worksheet = workbook.active
        data = []
        headers = [cell.value for cell in worksheet[1]]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return Response(data)


class SMOrderQueue(APIView):

    def get(self, request, format=None):
        queryset = models.SMOrderQueue.objects.all()
        serializer_class = serializers.SMOrderQueueSerializer(
            queryset, many=True)
        return Response(serializer_class.data)

    def post(self, request):
        data = json.loads(request.body)
        try:
            models.SMOrderQueue.objects.create(
                **data
            )
        except IntegrityError as e:
            return Response({'status': 'failed', 'reason': e}, status=HTTP_400_BAD_REQUEST)
        order_id = models.SMOrderQueue.objects.latest('id').id
        return Response({'status': 'success', 'id': order_id}, status=HTTP_201_CREATED)


class SMUpdateOrderQueue(APIView):
    def put(self, request, id):
        data = json.loads(request.body)
        try:
            models.SMOrderQueue.objects.filter(id=id).update(
                **data
            )
        except Exception as e:
            return Response({'status': 'failed', 'message': e}, status=HTTP_400_BAD_REQUEST)
        return Response({'status': 'success'}, status=HTTP_200_OK)

    def delete(self, request, id):
        try:
            models.SMOrderQueue.objects.filter(id=id).delete()
        except Exception as e:
            return Response({'status': 'failed', 'message': e}, status=HTTP_400_BAD_REQUEST)
        return Response({'status': 'success'}, status=HTTP_200_OK)


class PenCalculateDistance(APIView):
    def post(self, request):
        log("Penészmentesítés MiniCRM webhook meghívva",
            "INFO", "pen_calculate_distance")
        data = json.loads(str(request.body)[2:-1])["Data"]
        telephely = "Budapest, Nagytétényi út 218, 1225"

        address = f"{data['Cim2']} {data['Telepules']}, {data['Iranyitoszam']} {data['Orszag']}"
        gmaps_result = calculate_distance(
            start=telephely, end=codecs.unicode_escape_decode(address)[0])
        if gmaps_result == "Error":
            log("Penészmentesítés MiniCRM webhook sikertelen", "ERROR", "pen_calculate_distance",
                f"Hiba a Google Maps API-al való kommunikáció során {address}, adatlap id: {data['Id']}")
            return Response({'status': 'error'}, status=HTTP_200_OK)
        duration = gmaps_result["duration"] / 60
        distance = gmaps_result["distance"] // 1000
        formatted_duration = f"{math.floor(duration//60)} óra {math.floor(duration%60)} perc"
        fee_map = {
            0: 20000,
            31: 25000,
            101: 30000,
            201: 35000,
        }
        fee = fee_map[[i for i in fee_map.keys() if i < distance][-1]]

        try:
            get_street_view(location=codecs.unicode_escape_decode(address)[0])
        except Exception as e:
            log("Penészmentesítés MiniCRM webhook sikertelen", "FAILED", e)
        street_view_url = get_street_view_url(
            location=codecs.unicode_escape_decode(address)[0])
        response = update_adatlap_fields(data["Id"], {
            "IngatlanKepe": "https://www.dataupload.xyz/static/images/google_street_view/street_view.jpg", "UtazasiIdoKozponttol": formatted_duration, "Tavolsag": distance, "FelmeresiDij": fee, "StreetViewUrl": street_view_url, "BruttoFelmeresiDij": round(fee*1.27), "UtvonalAKozponttol": f"https://www.google.com/maps/dir/?api=1&origin=M%C3%A1tra+u.+17,+Budapest,+1224&destination={codecs.decode(address, 'unicode_escape')}&travelmode=driving"})
        if response.code == 200:
            log("Penészmentesítés MiniCRM webhook sikeresen lefutott",
                "SUCCESS", "pen_calculate_distance")
        else:
            log("Penészmentesítés MiniCRM webhook sikertelen",
                "ERROR", "pen_calculate_distance", response.reason)
        return Response({'status': 'success'}, status=HTTP_200_OK)


class Deploy(APIView):
    def post(self, request):
        try:
            subprocess.Popen(
                ["git", "pull"])
            subprocess.Popen(
                "pkill -TERM $(ps -C gunicorn -o pid= | head -n 1)", shell=True)
            return Response("Successfully deployed", status=HTTP_200_OK)
        except:
            return Response(f"Error while deploying", status=HTTP_500_INTERNAL_SERVER_ERROR)
